#coding:utf-8


import json
from mtrops import settings
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pymysql import connect
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font,Color,colors,GradientFill,NamedStyle



def ConnetMysql(sql):
    try:
        db = connect("127.0.0.1", "root", "mysql", "mtrops", charset="utf8")
        cursor = db.cursor()
        cursor.execute(sql)
        data = cursor.fetchall()
        db.commit()
        db.close()
        return  data
    except:
        db.rollback()


def GetData(sql):



    data = ConnetMysql(sql)


    op_info = ['李祖祥','13432089040','13432089040@139.com']


    data_list = []
    for host in data:
        n = 1
        hosts = list(host)+op_info
        total = 0
        host_info = []
        for i in  hosts:


            if n == 14 or n == 15:
                try:
                    i = str(round(float(i)/1024/1024,2)) + ' GB'
                except:
                    i = 'Null'

            if n == 18 or  n ==19 or n == 20:
                try:
                    total+= json.loads(i.replace("'",'"'))['total']/1024/1024/1024
                except:
                    total +=0

                if n == 20:
                    host_info.append(str(total)+' GB')
                n += 1
            else:
                host_info.append(i)
                n += 1

        data_list.append(host_info)

    return data_list



def headStyle():
    ft = Font(size=14,name='SimSun')  # color="0F0F0F"字体颜色,italic=False,是否斜体,字体样式,大小,bold=False是否粗体

    align = Alignment(horizontal="center", vertical="center")  # 居中对齐

    fill = PatternFill("solid", fgColor="FFFF00")  # 背景填充颜色

    # fill = GradientFill(stop=("000000", "FFFFFF")) #背景填充颜色，渐变

    bd = Side(style='thin', color="000000")

    border = Border(left=bd, top=bd, right=bd, bottom=bd)

    mystyle = NamedStyle(name="mystyle")

    mystyle.font = ft
    mystyle.alignment = align
    mystyle.fill = fill
    mystyle.border = border

    return mystyle


def boderStyle():
    ft = Font(size=12,name='SimSun')  # color="0F0F0F"字体颜色,italic=False,是否斜体,字体样式,大小,bold=False是否粗体

    align = Alignment(horizontal="center", vertical="center")  # 居中对齐

    #fill = PatternFill("solid", fgColor="FFFF00")  # 背景填充颜色

    # fill = GradientFill(stop=("000000", "FFFFFF")) #背景填充颜色，渐变

    bd = Side(style='thin', color="000000")

    border = Border(left=bd, top=bd, right=bd, bottom=bd)

    bdstyle = NamedStyle(name="bdstyle")

    bdstyle.font = ft
    bdstyle.alignment = align
   # mystyle.fill = fill
   # mystyle.border = border

    return bdstyle



def makeExcel(data_list):


    base_dir = settings.BASE_DIR

    path_dir = base_dir + '/static/media/'

    wb = Workbook()

    filename = path_dir + 'cmdb.xlsx'

    ws = wb.active

    ws.title = u'资产清单'

    N= 1
    head = ['IP地址','主机名','机房','设备类型','主机组','虚拟设备','管理用户','用户密码','描述','远程端口','系统类型','内核版本','系统版本','物理内存','Swap虚拟内存','CPU类型','CPU核数','磁盘空间','开机状态','管理员','电话','邮箱']

    data_list.insert(0,head)

    head_Style = headStyle()
    boder_Style = boderStyle()
    for i in data_list:
        n =1

        for j in i:
            col = get_column_letter(n)
            ws["%s%d" % (col,N)] = j

            if N == 1:
                ws["%s%d" % (col, N)].style = head_Style
                ws.column_dimensions['%s' % col ].width = 18  # 设置列宽
            else:
                ws["%s%d" % (col, N)].style = boder_Style

            n+=1
        N+=1

    ws.row_dimensions[1].height = 30  #设置行高


    wb.save(filename)
    

def main(sql):
    data_list = GetData(sql)
    makeExcel(data_list)



if __name__ == '__main__':
    sql = '''
        select a.IP,a.hostname,b.idc_name,a.device_type,c.group_name,a.service_type,a.username,a.passwd,a.msg,a.port,a.os_type,a.kernel,a.os_version,a.mem_total,a.mem_swap,a.cpu_type,a.cpu_num,a.mount_home ,a.mount_root,a.mount_alidata,a.status from app_cmdb_hostinfo a,app_cmdb_idc b,app_cmdb_hostgroup c where a.idc_id=b.id and a.host_group_id=c.id
        '''
    main(sql)
